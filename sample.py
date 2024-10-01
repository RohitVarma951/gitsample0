# print("Hello")
# name = 'John Smith'
# age = 20
# is_new = True

# name = input('what is your name ? ')
# color = input('what is your favourite color ? ')
# print(name + ' likes ' + color)

# weight_in_kgs = float(input('weight in kgs : '))
# weight_in_pounds = weight_in_kgs * 2.20462
# print(f'weight in pounds : {weight_in_pounds}')

# c = 'dsedfsvffdsv fvdfvfdfdfv'
# print(c.capitalize())

# x = 10
# x += 3
# print(x)

# name = 'djieojdewhfiuefjdnewjfdiwejniufhewiufheiohfiuwheufdjiowesehifweiofek'
# l = len(name)
# if l<3:
#     print('name must be atleast 3 characters')
# elif l>50:
#     print('name can be a maximum of 50 characters')
# else:
#     print('name looks good!')

# w = float(input('weight : '))
# k = input('(L)bs or (K)gs :')
# if k.upper() == 'L' :
#     w *= 0.453592
#     print(f'You are {w}Kilos')
# elif k.upper() == 'K' :
#     w *= 2.20462
#     print(f'You are {w}Pounds')
# else:
#     print('wrong metrics')

# secret = 9
# i = 0
# while i<3 :
#     inp = int(input('Guess : '))
#     i += 1
#     if inp == secret :
#         print('You Won!')
#         break
# else:
#     print('You Failed!')

# msg = ''
# is_started = False
# while True:
#     msg = input('>').lower()
#     if msg == 'help':
#         print('start - to start the car')
#         print('stop  - to stop the car')
#         print('quit  - to exit')
#     elif msg == 'start':
#         if is_started :
#             print('Car is already started!')
#         else :
#             is_started = True
#             print('Car started...Ready to go!')
#     elif msg == 'stop':
#         if is_started :
#             is_started = False
#             print('Car Stopped!')
#         else :
#             print('Car is already stopped!')
#     elif msg == 'quit':
#         break
#     else:
#         print("I don't understand that...")
#     prev = msg

# prices = [10,20,30]
# total = 0
# for i in prices:
#     total += i
# print(f'total : {total}')

# l = [5,2,5,2,2]
# for i in l:
#     for j in range(i):
#         print('x',end='')
#     print()

# l = [1,2,25,35,634,5,543,56,4]
# max = l[0]
# for i in l[1:]:
#     if i > max:
#         max = i
# print(max)

# l = [1,2,3,4,5,1,2,3,4,5]
# l.remove(2)
# print(l.count(3))
# l.sort()
# print(l)
# l.reverse()
# print(l)

# l = [1,2,3,4,5,1,2,3,4,5]
# for i in l:
#     if l.count(i) > 1 :
#         l.remove(i)
# print(l)

# num_dict = {
#     '1' : 'One',
#     '2' : 'Two',
#     '3' : 'Three',
#     '4' : 'Four'
# }
# s = input('Phone : ')
# out = ''
# for i in s:
#     out += num_dict.get(i,'!') + ' '
# print(out)

# def emote(inp):
#     w = inp.split(' ')
#     emojis ={
#         ':)' : '😀',
#         ':(' : '😞'
#     }
#     out = ''
#     for i in w:
#         out += emojis.get(i,i) + ' '
#     return out


# inp = input('>')
# out = emote(inp)
# print(out)

# class Person():
#     def __init__(self,name):
#         self.name = name
    
#     def talk(self):
#         print(f'{self.name} is talking.')


# p1 = Person('John')
# print(p1.name)
# p1.talk()

# p2 = Person('Bob')
# print(p2.name)
# p2.talk()

# import utils
# numbers = [314,42,54,5434,45325,43,4532,34532,2343]
# print(utils.find_max(numbers))

# import random

# class Dice:
#     def roll(self):
#         return random.randint(1,6), random.randint(1,6)

# d = Dice()
# print(d.roll())


# exists | mkdir | rmdir | glob
# from pathlib import Path

# path = Path()
# for file in path.glob('*'):
#     print(file)


import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = cell.value * 0.9

    values = Reference(sheet,
                        min_row = 2,
                        max_row = sheet.max_row,
                        min_col = 4,
                        max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a6')

    wb.save(filename)